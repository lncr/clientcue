from datetime import datetime
from django.conf import settings
from openpyxl import Workbook
from openpyxl import load_workbook

from applications.contacts.models import Contact
from applications.tags.models import Tag


def from_db_to_excel(contacts, user):
    """
    Writes contacts information into xlsx file
    Returns path to generated file
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f'Contacts of user {user.first_name} {user.last_name}'
    for row in range(1, contacts.count()+1):
        contact = contacts[row-1]
        worksheet.cell(column=1, row=row, value=f'{contact.name}')
        worksheet.cell(column=2, row=row, value=f'{contact.phone_number}')
        worksheet.cell(column=3, row=row, value=f'{contact.created_at}')
        tag_names_list = [tag.name for tag in contact.tags.all()]
        tag_names = ', '.join(tag_names_list)
        worksheet.cell(column=4, row=row, value=tag_names)
        worksheet.cell(column=5, row=row, value=f'{contact.rating}')
    filepath = f'{user.first_name}{user.last_name}.xlsx'
    filename = settings.MEDIA_ROOT + '/' + filepath
    workbook.save(filename=filename)
    return filepath


def from_excel_to_db(path, user):
    """
    Reads xlsx file and writes data to db
    """
    workbook = load_workbook(settings.MEDIA_ROOT + '/' + path)
    for worksheet in workbook.worksheets:
        for row in worksheet.rows:
            try:
                contact = Contact.objects.create(
                    name=row[0].value,
                    phone_number=row[1].value,
                    created_at=datetime.strptime(row[2].value, '%Y-%m-%d'),
                    rating=row[4].value
                )
                tag_names = row[3].split(', ')
                for tag_name in tag_names:
                    tag = Tag.objects.get_or_create(defaults={'owner_id': user.id, 'name': tag_name})
                    contact.tags.add(tag)
            except:
                continue
